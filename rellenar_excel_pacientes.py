import os
import pandas as pd
from openpyxl import Workbook

# Ruta base donde están los pacientes
base_dir = "/Users/claudiacastrillonalvarez/Desktop/patients_tiff_104_107"

# Lista de pacientes que quieres procesar
pacientes = [f"PACIENTE_{i}" for i in range(104, 108)]

# Modalidades de imagen
tipos = ["MRC", "PEI"]

# Ruta donde guardar los excels
output_dir = "/Users/claudiacastrillonalvarez/Desktop"

# Procesar cada modalidad por separado
for tipo in tipos:
    wb = Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía por defecto

    for paciente in pacientes:
        ruta_tipo = os.path.join(base_dir, paciente, tipo)

        if not os.path.isdir(ruta_tipo):
            print(f"⚠️ No se encontró la carpeta: {ruta_tipo}")
            continue

        archivos = sorted([
            f for f in os.listdir(ruta_tipo)
            if os.path.isfile(os.path.join(ruta_tipo, f)) and f.lower().endswith(('.tif', '.tiff'))
        ])

        # Crear DataFrame
        df = pd.DataFrame({
            "File Name": archivos,
            "Annotation": ["" for _ in archivos]
        })

        # Crear hoja con nombre del paciente (formato: PACIENTE 104)
        hoja = paciente.replace("_", " ")
        ws = wb.create_sheet(title=hoja)

        # Escribir encabezados
        for col_idx, column in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=column)

        # Escribir datos
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        print(f"✅ Añadida hoja: {hoja} ({len(archivos)} imágenes)")

    # Guardar archivo Excel por modalidad
    output_file = os.path.join(output_dir, f"{tipo}_TIFF_Annotations.xlsx")
    wb.save(output_file)
    print(f"💾 Guardado: {output_file}")
